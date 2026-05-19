from __future__ import annotations

import json
from functools import lru_cache
from pathlib import Path

import openpyxl

from .config import get_settings
from .schemas import (
    AnalysisSideSource,
    BasicParameters,
    BoltStrengthRow,
    ColumnScenarioInput,
    CornerBraceRow,
    FoundationSoilLayer,
    ProjectMetadata,
    ProjectState,
    ReferenceData,
    SectionProperty,
    SoilLayer,
    SupportRow,
    WaleRow,
    BraceRow,
)


def _workbooks() -> tuple[openpyxl.Workbook, openpyxl.Workbook]:
    settings = get_settings()
    if not settings.workbook_path.exists():
        raise FileNotFoundError(f"找不到 Excel 參考檔：{settings.workbook_path}")
    formula_wb = openpyxl.load_workbook(
        settings.workbook_path, read_only=True, data_only=False, keep_vba=True
    )
    value_wb = openpyxl.load_workbook(
        settings.workbook_path, read_only=True, data_only=True, keep_vba=True
    )
    return formula_wb, value_wb


def _normalize_section_name(name: str | None) -> str:
    if not name:
        return ""
    normalized = (
        str(name)
        .upper()
        .replace("×", "X")
        .replace("*", "X")
        .replace(" ", "")
        .replace("－", "-")
    )
    return normalized


def _reference_override_path() -> Path:
    return get_settings().reference_overrides_path


def _sanitize_reference_data(reference_data: ReferenceData) -> ReferenceData:
    normalized_sections: list[SectionProperty] = []
    seen_section_names: set[str] = set()
    for section in reference_data.sections:
        name = _normalize_section_name(section.name)
        if not name:
            raise ValueError("型鋼名稱不可空白。")
        if name in seen_section_names:
            raise ValueError(f"型鋼名稱重複：{name}")
        seen_section_names.add(name)
        normalized_sections.append(section.model_copy(update={"name": name}))

    normalized_bolts: list[BoltStrengthRow] = []
    for bolt in reference_data.bolts:
        grade = str(bolt.grade or "").strip()
        if not grade:
            raise ValueError("螺栓等級不可空白。")
        sizes = {
            str(size).strip(): float(value)
            for size, value in bolt.sizes.items()
            if str(size).strip()
        }
        normalized_bolts.append(
            bolt.model_copy(
                update={
                    "grade": grade,
                    "sizes": sizes,
                }
            )
        )

    return reference_data.model_copy(
        update={
            "sections": normalized_sections,
            "bolts": normalized_bolts,
        }
    )


def clear_reference_cache() -> None:
    load_reference_data.cache_clear()


def save_reference_data(reference_data: ReferenceData) -> ReferenceData:
    sanitized = _sanitize_reference_data(reference_data)
    override_path = _reference_override_path()
    override_path.parent.mkdir(parents=True, exist_ok=True)
    override_path.write_text(sanitized.model_dump_json(indent=2), encoding="utf-8")
    clear_reference_cache()
    return load_reference_data()


def reset_reference_overrides() -> ReferenceData:
    override_path = _reference_override_path()
    if override_path.exists():
        override_path.unlink()
    clear_reference_cache()
    return load_reference_data()


@lru_cache(maxsize=1)
def load_reference_data() -> ReferenceData:
    override_path = _reference_override_path()
    if override_path.exists():
        payload = json.loads(override_path.read_text(encoding="utf-8"))
        return _sanitize_reference_data(ReferenceData.model_validate(payload))

    _, value_wb = _workbooks()
    sections = _load_sections(value_wb)
    bolts = _load_bolts(value_wb)
    basic_defaults = _load_basic_defaults(value_wb)
    return ReferenceData(sections=sections, bolts=bolts, basic_defaults=basic_defaults)


def _load_sections(wb: openpyxl.Workbook) -> list[SectionProperty]:
    ws = wb["H"]
    sections: list[SectionProperty] = []
    for row in ws.iter_rows(min_row=3, max_col=19, values_only=True):
        if not row[1]:
            continue
        name = _normalize_section_name(row[1])
        if not name or row[3] is None:
            continue
        sections.append(
            SectionProperty(
                name=name,
                depth_cm=float(row[3]) / 10.0,
                flange_width_cm=float(row[4]) / 10.0,
                web_thickness_cm=float(row[5]) / 10.0,
                flange_thickness_cm=float(row[6]) / 10.0,
                area_cm2=float(row[8]),
                unit_weight_kgf_per_m=float(row[9]),
                ix_cm4=float(row[10]),
                iy_cm4=float(row[11]),
                rx_cm=float(row[12]),
                ry_cm=float(row[13]),
                rt_cm=float(row[14]),
                sx_cm3=float(row[15]),
                sy_cm3=float(row[16]),
                zx_cm3=float(row[17]),
                zy_cm3=float(row[18]),
            )
        )
    return sections


def _load_bolts(wb: openpyxl.Workbook) -> list[BoltStrengthRow]:
    ws = wb["bolt"]
    tension_header = [ws.cell(4, col).value for col in range(4, 11)]
    shear_header = [ws.cell(12, col).value for col in range(4, 11)]
    bolts: list[BoltStrengthRow] = []
    for row in range(5, 8):
        sizes = {
            str(tension_header[idx]): float(ws.cell(row, 4 + idx).value)
            for idx in range(len(tension_header))
            if tension_header[idx]
        }
        bolts.append(
            BoltStrengthRow(
                grade=str(ws.cell(row, 2).value),
                ft_tf_per_cm2=float(ws.cell(row, 3).value),
                sizes=sizes,
            )
        )
    for row in range(13, 16):
        sizes = {
            str(shear_header[idx]): float(ws.cell(row, 4 + idx).value)
            for idx in range(len(shear_header))
            if shear_header[idx]
        }
        bolts.append(
            BoltStrengthRow(
                grade=f"{ws.cell(row, 2).value} (剪力)",
                fv_tf_per_cm2=float(ws.cell(row, 3).value),
                sizes=sizes,
            )
        )
    return bolts


def _load_basic_defaults(wb: openpyxl.Workbook) -> BasicParameters:
    ws = wb["基本參數"]
    return BasicParameters(
        e_tf_per_cm2=float(ws["C15"].value),
        fy_tf_per_cm2=float(ws["C16"].value),
        cm_factor=float(ws["C18"].value),
        surcharge_wl_tf_per_m=float(ws["C19"].value),
        alpha_support=float(ws["C20"].value),
        alpha_wale=float(ws["C21"].value),
        alpha_brace=float(ws["C22"].value),
        alpha_column=float(ws["C23"].value),
        psi_material=float(ws["C24"].value),
        wall_type=str(ws["C77"].value),
        wall_thickness_cm=float(ws["D78"].value),
        wall_fc_kg_per_cm2=float(ws["D79"].value),
    )


def find_section(name: str) -> SectionProperty:
    normalized = _normalize_section_name(name)
    sections = load_reference_data().sections
    for section in sections:
        if section.name == normalized:
            return section
    if normalized.startswith("H"):
        alt = "R" + normalized
        for section in sections:
            if section.name == alt:
                return section
    if normalized.startswith("RH"):
        alt = normalized[1:]
        for section in sections:
            if section.name == alt:
                return section
    raise KeyError(f"找不到型鋼資料：{name}")


@lru_cache(maxsize=1)
def load_default_project() -> ProjectState:
    _, wb = _workbooks()
    design_ws = wb["各項設計參數"]
    middle_ws = wb["中間柱"]
    normal_ws = wb["共構柱 (一般)"]
    crane_ws = wb["共構柱 (大吊車)"]

    top_supports = [
        SupportRow(
            level_label=str(design_ws.cell(row, 1).value),
            support_count=int(design_ws.cell(row, 2).value or 0),
            section_name=_normalize_section_name(design_ws.cell(row, 3).value),
            axial_force_t=float(design_ws.cell(row, 4).value or 0),
            temp_force_t=float(design_ws.cell(row, 5).value or 0),
            spacing_m=float(design_ws.cell(row, 6).value or 0),
        )
        for row in range(5, 13)
        if design_ws.cell(row, 3).value
    ]
    bottom_supports = [
        SupportRow(
            level_label=str(design_ws.cell(row, 1).value),
            support_count=int(design_ws.cell(row, 2).value or 0),
            section_name=_normalize_section_name(design_ws.cell(row, 3).value),
            axial_force_t=float(design_ws.cell(row, 4).value or 0),
            temp_force_t=float(design_ws.cell(row, 5).value or 0),
            spacing_m=float(design_ws.cell(row, 6).value or 0),
        )
        for row in range(19, 27)
        if design_ws.cell(row, 3).value
    ]
    top_wales: list[WaleRow] = []
    top_braces: list[BraceRow] = []
    for row in range(33, 39):
        if design_ws.cell(row, 3).value:
            top_wales.append(
                WaleRow(
                    level_label=str(design_ws.cell(row, 1).value),
                    wale_count=int(design_ws.cell(row, 2).value or 0),
                    section_name=_normalize_section_name(design_ws.cell(row, 3).value),
                    span_m=float(design_ws.cell(row, 6).value or 0)
                    + float(design_ws.cell(row, 7).value or 0),
                    support_spacing_m=float(top_supports[row - 33].spacing_m if row - 33 < len(top_supports) else 0),
                    line_load_tf_per_m=0.0,
                )
            )
        if design_ws.cell(row, 5).value:
            top_braces.append(
                BraceRow(
                    level_label=str(design_ws.cell(row, 1).value),
                    section_name=_normalize_section_name(design_ws.cell(row, 5).value),
                    l1_m=float(design_ws.cell(row, 6).value or 0),
                    l2_m=float(design_ws.cell(row, 7).value or 0),
                    angle_deg=float(design_ws.cell(row, 8).value or 45),
                    tributary_line_load_tf_per_m=0.0,
                )
            )
    bottom_wales: list[WaleRow] = []
    bottom_braces: list[BraceRow] = []
    for row in range(48, 54):
        if design_ws.cell(row, 3).value:
            support_row = bottom_supports[row - 48] if row - 48 < len(bottom_supports) else None
            bottom_wales.append(
                WaleRow(
                    level_label=str(design_ws.cell(row, 1).value),
                    wale_count=int(design_ws.cell(row, 2).value or 0),
                    section_name=_normalize_section_name(design_ws.cell(row, 3).value),
                    span_m=float(design_ws.cell(row, 6).value or 0)
                    + float(design_ws.cell(row, 7).value or 0),
                    support_spacing_m=float(support_row.spacing_m if support_row else 0),
                    line_load_tf_per_m=0.0,
                )
            )
        if design_ws.cell(row, 5).value:
            bottom_braces.append(
                BraceRow(
                    level_label=str(design_ws.cell(row, 1).value),
                    section_name=_normalize_section_name(design_ws.cell(row, 5).value),
                    l1_m=float(design_ws.cell(row, 6).value or 0),
                    l2_m=float(design_ws.cell(row, 7).value or 0),
                    angle_deg=float(design_ws.cell(row, 8).value or 45),
                    tributary_line_load_tf_per_m=0.0,
                )
            )
    corner_ws = wb["大角撐"]
    corner_braces = [
        CornerBraceRow(
            level_label=str(corner_ws.cell(row, 1).value),
            section_name=_normalize_section_name(corner_ws.cell(row, 2).value),
            length_m=float(corner_ws.cell(row, 3).value or 0),
            axial_force_t=float(corner_ws.cell(row, 4).value or 0),
        )
        for row in range(5, 13)
        if corner_ws.cell(row, 2).value
    ]
    default_soils = _load_default_soils(wb)
    all_supports = [*top_supports, *bottom_supports]
    columns = [
        _column_from_sheet("中間柱", middle_ws, all_supports, default_soils, "middle"),
        _column_from_sheet("共構柱 (一般)", normal_ws, all_supports, default_soils, "composite_normal"),
        _column_from_sheet("共構柱 (大吊車)", crane_ws, all_supports, default_soils, "composite_crane"),
    ]
    basic_params = _load_basic_defaults(wb)
    _backfill_line_loads(top_supports, top_wales, top_braces, basic_params)
    _backfill_line_loads(bottom_supports, bottom_wales, bottom_braces, basic_params)
    return ProjectState(
        metadata=ProjectMetadata(name="Excel 轉換範例專案", location="本地工作區"),
        basic_parameters=basic_params,
        top_analysis_source=AnalysisSideSource(mode="manual"),
        bottom_analysis_source=AnalysisSideSource(mode="manual"),
        top_supports=top_supports,
        bottom_supports=bottom_supports,
        top_wales=top_wales,
        bottom_wales=bottom_wales,
        top_braces=top_braces,
        bottom_braces=bottom_braces,
        corner_braces=corner_braces,
        columns=columns,
    )


def _backfill_line_loads(
    supports: list[SupportRow],
    wales: list[WaleRow],
    braces: list[BraceRow],
    basic_params: BasicParameters,
) -> None:
    support_by_level = {row.level_label: row for row in supports}
    for wale in wales:
        support = support_by_level.get(wale.level_label)
        if not support:
            continue
        wale.line_load_tf_per_m = round(
            (support.axial_force_t + support.temp_force_t)
            * max(support.support_count, 1)
            / max(support.spacing_m, 1e-6),
            3,
        )
    wale_by_level = {row.level_label: row for row in wales}
    for brace in braces:
        wale = wale_by_level.get(brace.level_label)
        if wale:
            brace.tributary_line_load_tf_per_m = wale.line_load_tf_per_m


def _load_default_soils(wb: openpyxl.Workbook) -> list[FoundationSoilLayer]:
    ws = wb["Soil"]
    soils: list[FoundationSoilLayer] = []
    for row in range(3, 10):
        if ws.cell(row, 2).value is None:
            continue
        name = str(ws.cell(row, 3).value)
        soil_type = "clay" if name.upper().startswith("C") else "sand"
        soils.append(
            FoundationSoilLayer(
                index=int(ws.cell(row, 2).value),
                name=name,
                thickness_m=float(ws.cell(row, 4).value or 0),
                depth_m=float(ws.cell(row, 5).value or 0),
                n_value=float(ws.cell(row, 6).value or 0),
                su_t_per_m2=float(ws.cell(row, 11).value or 0)
                if ws.cell(row, 11).value not in (None, "－")
                else None,
                soil_type=soil_type,
            )
        )
    return soils


def _column_from_sheet(
    title: str,
    ws: openpyxl.worksheet.worksheet.Worksheet,
    support_rows: list[SupportRow],
    soil_layers: list[FoundationSoilLayer],
    variant: str,
) -> ColumnScenarioInput:
    size_x = float(ws["L3"].value or 0.8)
    size_y = float(ws["M3"].value or 2.5)
    return ColumnScenarioInput(
        title=title,
        variant=variant,  # type: ignore[arg-type]
        enabled=True,
        column_section_name=_normalize_section_name(ws["C2"].value),
        support_rows=support_rows,
        foundation_type=str(ws["C3"].value or "鑽掘或引孔樁"),
        foundation_shape=str(ws["K2"].value or "(直徑)"),
        foundation_size_x_m=size_x,
        foundation_size_y_m=size_y,
        column_length_m=float(ws["I2"].value or 20),
        kh_kg_per_cm3=4.0,
        bottom_to_excavation_distance_m=4.0,
        eccentricity_x_m=0.53,
        eccentricity_y_m=0.0,
        embedment_length_cm=300.0,
        concrete_strength_kg_per_cm2=float(ws["C67"].value or 175)
        if isinstance(ws["C67"].value, (int, float))
        else 175.0,
        soil_layers=soil_layers,
    )


def workbook_path() -> Path:
    return get_settings().workbook_path
